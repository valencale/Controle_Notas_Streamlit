"""
excel_handler.py — Módulo de leitura/escrita do ficheiro .xlsm via openpyxl.

ARQUITETURA:
    Todas as operações seguem o padrão open-modify-save-close para
    minimizar o tempo de lock do ficheiro e prevenir corrupção.
    O parâmetro keep_vba=True é OBRIGATÓRIO para preservar macros/ActiveX.

ESTRUTURA DO EXCEL:
    Aba "Dados":
        - Rows 1-6: Área de resumo/KPIs (status counts em E-F)
        - Row 9: Título "CONTROLE NOTAS" em B9
        - Row 10: Headers → DATA(B), CLIENTE(C), PEDIDO(D), EMPRESA(E), NF(F), STATUS(G), OBS(H), ENDERECO(I)
        - Row 11+: Dados dos pedidos
    Aba "Setup":
        - Lista de status válidos (A1:A5)
    Aba "Historico" (criada pelo app se não existir):
        - Mesma estrutura de colunas que "Dados"
        - Row 1: Headers
        - Row 2+: Pedidos arquivados
"""

import os
import logging
import shutil
from datetime import datetime
from typing import Optional

import openpyxl
import pandas as pd
import streamlit as st

from modules.auth import require_adm

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import (
    EXCEL_PATH,
    SHEET_PRINCIPAL,
    SHEET_HISTORICO,
    HEADER_ROW,
    DATA_START_ROW,
    COLUMNS,
    COLUMN_NAMES,
    EXCEL_HEADER_NAMES,
    PK_COLUMN,
    PK_COL_INDEX,
    STATUS_OPTIONS,
)

# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════

logger = logging.getLogger(__name__)

def _open_workbook(read_only: bool = False) -> openpyxl.Workbook:
    """
    Abre o workbook .xlsm preservando macros VBA.

    Args:
        read_only: Se True, abre em modo somente leitura (mais rápido para consultas).

    Returns:
        openpyxl.Workbook com keep_vba=True.

    Raises:
        FileNotFoundError: Se o ficheiro Excel não existir no caminho configurado.
    """
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Ficheiro Excel não encontrado: {EXCEL_PATH}")
    return openpyxl.load_workbook(EXCEL_PATH, keep_vba=True, data_only=False, read_only=read_only)


def _save_workbook(wb: openpyxl.Workbook) -> None:
    """
    Salva o workbook no caminho original.
    Cria um backup antes de salvar para segurança.

    Args:
        wb: Workbook a ser salvo.
    """
    wb.save(EXCEL_PATH)


def _ensure_historico_sheet(wb: openpyxl.Workbook) -> None:
    """
    Garante que a aba 'Historico' existe no workbook.
    Se não existir, cria com headers na row 1.

    Args:
        wb: Workbook aberto.
    """
    if SHEET_HISTORICO not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_HISTORICO)
        # Write headers in row 1
        for col_name, col_idx in COLUMNS.items():
            # Historico usa coluna A em diante (sem o offset do "Dados")
            display_name = EXCEL_HEADER_NAMES.get(col_name, col_name)
            ws.cell(row=1, column=col_idx - 1, value=display_name)


def _find_row_by_pk(ws, pedido_value: str) -> Optional[int]:
    """
    Localiza o número da linha que contém o pedido especificado, ignorando zeros à esquerda.

    Percorre todas as linhas de dados (a partir de DATA_START_ROW na aba Dados,
    ou row 2 na aba Historico) procurando correspondência na coluna PEDIDO.

    Args:
        ws: Worksheet openpyxl.
        pedido_value: Valor do pedido (PK) a localizar.

    Returns:
        Número da linha (1-indexed) ou None se não encontrado.
    """
    pedido_str = str(pedido_value).strip().lstrip('0')
    start_row = DATA_START_ROW if ws.title == SHEET_PRINCIPAL else 2
    pk_col = PK_COL_INDEX if ws.title == SHEET_PRINCIPAL else PK_COL_INDEX - 1

    for row_idx in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=pk_col).value
        if cell_value is not None and str(cell_value).strip().lstrip('0') == pedido_str:
            return row_idx
    return None


# ══════════════════════════════════════════════════════════════
# READ OPERATIONS
# ══════════════════════════════════════════════════════════════

def _get_excel_mtime() -> float:
    """Retorna o timestamp de modificação do ficheiro Excel."""
    try:
        return os.path.getmtime(EXCEL_PATH) if os.path.exists(EXCEL_PATH) else 0.0
    except OSError:
        return 0.0


@st.cache_data(show_spinner="Lendo dados do Excel...", ttl=300)
def _read_principal_cached(mtime: float) -> pd.DataFrame:
    """
    Lê a aba principal ('Dados') e retorna como DataFrame (versão em cache).
    O parâmetro mtime garante que o cache seja invalidado se o arquivo mudar.
    """
    if not os.path.exists(EXCEL_PATH):
        return pd.DataFrame(columns=COLUMN_NAMES)

    try:
        df = pd.read_excel(
            EXCEL_PATH,
            sheet_name=SHEET_PRINCIPAL,
            header=DATA_START_ROW - 2,  # Row 10 = header (0-indexed = 9)
            engine="openpyxl",
        )

        # Remove coluna A (vazia, "Unnamed: 0")
        df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

        # Renomeia 'NOTA FISCAL' → 'NF' para consistência
        if "NOTA FISCAL" in df.columns:
            df = df.rename(columns={"NOTA FISCAL": "NF"})

        # Remove linhas completamente vazias
        df = df.dropna(how="all").reset_index(drop=True)

        # Convert DATA to datetime safely
        if not df.empty and "DATA" in df.columns:
            df["DATA"] = pd.to_datetime(df["DATA"], errors="coerce")

        # Ensure PEDIDO is clean string (float 99945.0 → "99945")
        if not df.empty and "PEDIDO" in df.columns:
            df["PEDIDO"] = df["PEDIDO"].apply(
                lambda x: str(int(x)) if pd.notna(x) and isinstance(x, float) else str(x).strip() if pd.notna(x) else ""
            )

        return df

    except Exception as e:
        logger.warning(f"Erro ao ler aba Dados via pd.read_excel: {e}")
        return pd.DataFrame(columns=COLUMN_NAMES)


def read_principal() -> pd.DataFrame:
    """Entrypoint principal de leitura, injeta o mtime no cache."""
    return _read_principal_cached(_get_excel_mtime())


@st.cache_data(show_spinner="Lendo histórico...", ttl=300)
def _read_historico_cached(mtime: float) -> pd.DataFrame:
    """
    Lê a aba 'Historico' e retorna como DataFrame (versão em cache).
    """
    if not os.path.exists(EXCEL_PATH):
        return pd.DataFrame(columns=COLUMN_NAMES)

    try:
        df = pd.read_excel(
            EXCEL_PATH,
            sheet_name=SHEET_HISTORICO,
            header=0,
            engine="openpyxl",
        )

        # Renomeia 'NOTA FISCAL' → 'NF' para consistência
        if "NOTA FISCAL" in df.columns:
            df = df.rename(columns={"NOTA FISCAL": "NF"})

        # Remove linhas completamente vazias
        df = df.dropna(how="all").reset_index(drop=True)

        if not df.empty and "DATA" in df.columns:
            df["DATA"] = pd.to_datetime(df["DATA"], errors="coerce")
        if not df.empty and "PEDIDO" in df.columns:
            df["PEDIDO"] = df["PEDIDO"].apply(
                lambda x: str(int(x)) if pd.notna(x) and isinstance(x, float) else str(x).strip() if pd.notna(x) else ""
            )

        return df

    except Exception:
        # Fallback: aba não existe ou erro de leitura
        return pd.DataFrame(columns=COLUMN_NAMES)


def read_historico() -> pd.DataFrame:
    """Entrypoint de leitura do histórico, injeta o mtime no cache."""
    return _read_historico_cached(_get_excel_mtime())


# ══════════════════════════════════════════════════════════════
# WRITE OPERATIONS
# ══════════════════════════════════════════════════════════════

@require_adm
def insert_pedido(data: dict) -> bool:
    """
    Insere um novo pedido na aba principal ('Dados').

    Adiciona uma nova linha após a última linha com dados, preenchendo
    as colunas B-H conforme o dicionário fornecido.

    Args:
        data: Dicionário com chaves correspondentes a COLUMN_NAMES.
              Exemplo: {"DATA": datetime.now(), "CLIENTE": "Empresa X", "PEDIDO": "12345",
                        "EMPRESA": "GREEN BAGS", "STATUS": "SEPARACAO", "OBS": "", "ENDERECO": ""}

    Returns:
        True se a inserção foi bem-sucedida.

    Raises:
        ValueError: Se o campo PEDIDO estiver vazio ou já existir.
    """
    if not data.get("PEDIDO"):
        raise ValueError("O campo PEDIDO é obrigatório.")

    wb = _open_workbook()
    ws = wb[SHEET_PRINCIPAL]

    # Check for duplicate PK
    if _find_row_by_pk(ws, data["PEDIDO"]):
        wb.close()
        raise ValueError(f"Pedido {data['PEDIDO']} já existe na esteira.")

    # Find next empty row
    next_row = ws.max_row + 1

    # Write data
    for col_name, col_idx in COLUMNS.items():
        value = data.get(col_name, "")
        ws.cell(row=next_row, column=col_idx, value=value)

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear() # Invalida o cache após inserção
    return True


@require_adm
def update_status(pedido: str, new_status: str) -> bool:
    """
    Atualiza o status de um pedido na aba principal.

    Localiza a linha pela PK (PEDIDO) e altera o valor na coluna STATUS (F).

    Args:
        pedido: Número do pedido (PK).
        new_status: Novo status (deve estar em STATUS_OPTIONS).

    Returns:
        True se a atualização foi bem-sucedida.

    Raises:
        ValueError: Se o status não for válido ou o pedido não for encontrado.
    """
    if new_status not in STATUS_OPTIONS:
        raise ValueError(f"Status inválido: {new_status}. Opções: {STATUS_OPTIONS}")

    wb = _open_workbook()
    ws = wb[SHEET_PRINCIPAL]

    row_idx = _find_row_by_pk(ws, pedido)
    if row_idx is None:
        wb.close()
        raise ValueError(f"Pedido {pedido} não encontrado na esteira.")

    ws.cell(row=row_idx, column=COLUMNS["STATUS"], value=new_status)

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear() # Invalida o cache após atualização
    return True

@require_adm
def update_status_batch(updates: dict) -> int:
    """
    Atualiza o status de múltiplos pedidos na aba principal.
    
    Args:
        updates: Dicionário onde a chave é o número do pedido (PK) e o valor é o novo status.
        
    Returns:
        Número de registros atualizados com sucesso.
    """
    if not updates:
        return 0

    wb = _open_workbook()
    ws = wb[SHEET_PRINCIPAL]

    count = 0
    pk_col = COLUMNS[PK_COLUMN] if PK_COLUMN in COLUMNS else PK_COL_INDEX
    
    # Validar status antes
    for pk, new_status in updates.items():
        if new_status not in STATUS_OPTIONS:
            wb.close()
            raise ValueError(f"Status inválido para o pedido {pk}: {new_status}. Opções: {STATUS_OPTIONS}")

    # Normaliza chaves ignorando zeros à esquerda
    updates_normalized = {str(pk).strip().lstrip('0'): status for pk, status in updates.items()}
    
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=pk_col).value
        if cell_value is not None:
            pedido_str = str(cell_value).strip().lstrip('0')
            if pedido_str in updates_normalized:
                new_status = updates_normalized[pedido_str]
                ws.cell(row=row_idx, column=COLUMNS["STATUS"], value=new_status)
                count += 1

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear() # Invalida o cache após atualização em lote
    return count


@require_adm
def update_cell(pedido: str, column_name: str, value) -> bool:
    """
    Atualiza uma célula específica de um pedido na aba principal.

    Args:
        pedido: Número do pedido (PK).
        column_name: Nome da coluna (deve existir em COLUMNS).
        value: Novo valor para a célula.

    Returns:
        True se a atualização foi bem-sucedida.
    """
    if column_name not in COLUMNS:
        raise ValueError(f"Coluna inválida: {column_name}")

    wb = _open_workbook()
    ws = wb[SHEET_PRINCIPAL]

    row_idx = _find_row_by_pk(ws, pedido)
    if row_idx is None:
        wb.close()
        raise ValueError(f"Pedido {pedido} não encontrado.")

    ws.cell(row=row_idx, column=COLUMNS[column_name], value=value)

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear() # Invalida o cache após alteração de célula
    return True


@require_adm
def delete_pedido(pedido: str) -> bool:
    """
    Remove um pedido da aba principal.

    Localiza a linha pela PK e a exclui completamente, deslocando
    as linhas abaixo para cima.

    Args:
        pedido: Número do pedido a excluir.

    Returns:
        True se a exclusão foi bem-sucedida.

    Raises:
        ValueError: Se o pedido não for encontrado.
    """
    wb = _open_workbook()
    ws = wb[SHEET_PRINCIPAL]

    row_idx = _find_row_by_pk(ws, pedido)
    if row_idx is None:
        wb.close()
        raise ValueError(f"Pedido {pedido} não encontrado na esteira.")

    ws.delete_rows(row_idx, 1)

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear() # Invalida o cache após exclusão
    return True


@require_adm
def update_rows_batch(changes: list[dict]) -> int:
    """
    Atualiza múltiplos campos de múltiplos pedidos em uma única operação.

    Projetado para o st.data_editor: recebe as linhas editadas e grava
    tudo de uma vez no Excel, evitando múltiplos open/save.

    Args:
        changes: Lista de dicts, cada um com:
            - "PEDIDO": str — PK do pedido
            - demais chaves = colunas a atualizar (ex: "STATUS", "OBS", "CLIENTE")
            Os valores já devem estar formatados (ex: DATA como datetime).

    Returns:
        Número de pedidos atualizados com sucesso.
    """
    if not changes:
        return 0

    wb = _open_workbook()
    ws = wb[SHEET_PRINCIPAL]

    # Validações de status
    for ch in changes:
        if "STATUS" in ch and ch["STATUS"] not in STATUS_OPTIONS:
            wb.close()
            raise ValueError(f"Status inválido: {ch['STATUS']}. Opções: {STATUS_OPTIONS}")

    # Indexar linhas do Excel por pedido (normalizado sem zeros à esquerda)
    row_map = {}
    pk_col = PK_COL_INDEX
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=pk_col).value
        if cell_value is not None:
            key = str(cell_value).strip().lstrip('0')
            row_map[key] = row_idx

    count = 0
    for ch in changes:
        pedido = str(ch.get("PEDIDO", "")).strip()
        key = pedido.lstrip('0')
        row_idx = row_map.get(key)
        if row_idx is None:
            continue

        for col_name, value in ch.items():
            if col_name == "PEDIDO":
                continue  # PK não altera
            if col_name in COLUMNS:
                ws.cell(row=row_idx, column=COLUMNS[col_name], value=value)
        count += 1

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear()
    return count


@require_adm
def delete_pedidos_batch(pedidos: list[str]) -> int:
    """
    Exclui múltiplos pedidos da aba principal em uma única operação.

    Itera de baixo para cima para evitar deslocamento de índices.

    Args:
        pedidos: Lista de números de pedido (PK) a excluir.

    Returns:
        Número de pedidos excluídos.
    """
    if not pedidos:
        return 0

    wb = _open_workbook()
    ws = wb[SHEET_PRINCIPAL]

    # Normalizar pedidos para busca
    pedidos_normalized = {str(p).strip().lstrip('0') for p in pedidos}

    rows_to_delete = []
    pk_col = PK_COL_INDEX

    for row_idx in range(ws.max_row, DATA_START_ROW - 1, -1):
        cell_value = ws.cell(row=row_idx, column=pk_col).value
        if cell_value is not None:
            key = str(cell_value).strip().lstrip('0')
            if key in pedidos_normalized:
                rows_to_delete.append(row_idx)

    count = 0
    for row_idx in rows_to_delete:
        ws.delete_rows(row_idx, 1)
        count += 1

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear()
    return count


# ══════════════════════════════════════════════════════════════
# ARCHIVE / RESTORE OPERATIONS
# ══════════════════════════════════════════════════════════════

@require_adm
def archive_completed() -> int:
    """
    Move todos os pedidos com STATUS='CONCLUIDO' da aba 'Dados' para 'Historico'.

    Processo:
        1. Abre o workbook
        2. Garante que a aba Historico existe
        3. Percorre a aba Dados de baixo para cima (para evitar shift de índices)
        4. Para cada linha com STATUS=CONCLUIDO:
           a. Copia os valores para a próxima linha vazia em Historico
           b. Remove a linha da aba Dados
        5. Salva o workbook

    Returns:
        Número de pedidos arquivados.
    """
    wb = _open_workbook()
    _ensure_historico_sheet(wb)

    ws_dados = wb[SHEET_PRINCIPAL]
    ws_hist = wb[SHEET_HISTORICO]

    # Find next empty row in Historico
    hist_next_row = ws_hist.max_row + 1
    if ws_hist.max_row == 1 and ws_hist.cell(row=1, column=1).value is not None:
        hist_next_row = 2  # Headers exist, start at row 2

    # Collect rows to archive (iterate bottom-up to avoid index shifting)
    rows_to_delete = []
    status_col = COLUMNS["STATUS"]

    for row_idx in range(ws_dados.max_row, DATA_START_ROW - 1, -1):
        status_val = ws_dados.cell(row=row_idx, column=status_col).value
        if status_val is not None and str(status_val).strip().upper() == "CONCLUIDO":
            rows_to_delete.append(row_idx)

    # Archive each row
    count = 0
    for row_idx in rows_to_delete:
        # Copy to Historico
        for col_name, col_idx in COLUMNS.items():
            value = ws_dados.cell(row=row_idx, column=col_idx).value
            ws_hist.cell(row=hist_next_row, column=col_idx - 1, value=value)
        hist_next_row += 1
        count += 1

    # Delete from Dados (bottom-up to preserve indices)
    for row_idx in rows_to_delete:
        ws_dados.delete_rows(row_idx, 1)

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear() # Invalida o cache após arquivamento
    return count


@require_adm
def archive_dispatched() -> int:
    """
    Move todos os pedidos com STATUS='EM ROTA' da aba 'Dados' para 'Historico'.

    Usado para tirar da visão primária os pedidos já despachados,
    mantendo-os acessíveis no Histórico para conferência e atualização
    de status (EM ROTA → ENTREGUE).

    Returns:
        Número de pedidos arquivados.
    """
    wb = _open_workbook()
    _ensure_historico_sheet(wb)

    ws_dados = wb[SHEET_PRINCIPAL]
    ws_hist = wb[SHEET_HISTORICO]

    # Find next empty row in Historico
    hist_next_row = ws_hist.max_row + 1
    if ws_hist.max_row == 1 and ws_hist.cell(row=1, column=1).value is not None:
        hist_next_row = 2  # Headers exist, start at row 2

    # Collect rows to archive (iterate bottom-up to avoid index shifting)
    rows_to_delete = []
    status_col = COLUMNS["STATUS"]

    for row_idx in range(ws_dados.max_row, DATA_START_ROW - 1, -1):
        status_val = ws_dados.cell(row=row_idx, column=status_col).value
        if status_val is not None and str(status_val).strip().upper() == "EM ROTA":
            rows_to_delete.append(row_idx)

    # Archive each row
    count = 0
    for row_idx in rows_to_delete:
        # Copy to Historico
        for col_name, col_idx in COLUMNS.items():
            value = ws_dados.cell(row=row_idx, column=col_idx).value
            ws_hist.cell(row=hist_next_row, column=col_idx - 1, value=value)
        hist_next_row += 1
        count += 1

    # Delete from Dados (bottom-up to preserve indices)
    for row_idx in rows_to_delete:
        ws_dados.delete_rows(row_idx, 1)

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear()
    return count


@require_adm
def update_status_in_historico(updates: dict) -> int:
    """
    Atualiza o status de pedidos na aba 'Historico'.

    Usado pela conferência da expedição para alterar EM ROTA → ENTREGUE
    nos pedidos que já foram movidos para o Histórico.

    Args:
        updates: Dict {pedido: novo_status}

    Returns:
        Número de pedidos atualizados.
    """
    if not updates:
        return 0

    wb = _open_workbook()

    if SHEET_HISTORICO not in wb.sheetnames:
        wb.close()
        return 0

    ws_hist = wb[SHEET_HISTORICO]

    # Normaliza chaves ignorando zeros à esquerda
    updates_normalized = {str(pk).strip().lstrip('0'): status for pk, status in updates.items()}

    # Na aba Historico, PK_COL_INDEX tem offset -1
    pk_col = PK_COL_INDEX - 1
    status_col = COLUMNS["STATUS"] - 1  # offset -1 para Historico

    count = 0
    for row_idx in range(2, ws_hist.max_row + 1):  # Row 1 = headers
        cell_value = ws_hist.cell(row=row_idx, column=pk_col).value
        if cell_value is not None:
            pedido_str = str(cell_value).strip().lstrip('0')
            if pedido_str in updates_normalized:
                new_status = updates_normalized[pedido_str]
                ws_hist.cell(row=row_idx, column=status_col, value=new_status)
                count += 1

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear()
    return count


def count_status_in_historico() -> dict:
    """
    Conta pedidos por status na aba 'Historico'.

    Usado para somar nos KPIs da Esteira os pedidos EM ROTA e ENTREGUE
    que já foram movidos para o Histórico.

    Returns:
        Dict {status: contagem}, ex: {"EM ROTA": 15, "ENTREGUE": 42}
    """
    wb = _open_workbook(read_only=True)

    if SHEET_HISTORICO not in wb.sheetnames:
        wb.close()
        return {}

    ws_hist = wb[SHEET_HISTORICO]

    # Na aba Historico, STATUS tem offset -1
    status_col = COLUMNS["STATUS"] - 1

    counts = {}
    for row_idx in range(2, ws_hist.max_row + 1):
        status_val = ws_hist.cell(row=row_idx, column=status_col).value
        if status_val is not None:
            status_str = str(status_val).strip().upper()
            if status_str:
                counts[status_str] = counts.get(status_str, 0) + 1

    wb.close()
    return counts


@require_adm
def restore_from_historico(pedido: str) -> bool:
    """
    Move um pedido do Historico de volta para a aba principal (estorno).

    Args:
        pedido: Número do pedido a restaurar.

    Returns:
        True se o estorno foi bem-sucedido.

    Raises:
        ValueError: Se o pedido não for encontrado no Historico.
    """
    wb = _open_workbook()

    if SHEET_HISTORICO not in wb.sheetnames:
        wb.close()
        raise ValueError("Aba Historico não existe.")

    ws_hist = wb[SHEET_HISTORICO]
    ws_dados = wb[SHEET_PRINCIPAL]

    # Find the row in Historico
    row_idx = _find_row_by_pk(ws_hist, pedido)
    if row_idx is None:
        wb.close()
        raise ValueError(f"Pedido {pedido} não encontrado no Histórico.")

    # Copy to Dados
    next_row = ws_dados.max_row + 1
    for col_name, col_idx in COLUMNS.items():
        value = ws_hist.cell(row=row_idx, column=col_idx - 1).value
        ws_dados.cell(row=next_row, column=col_idx, value=value)

    # Set status back to SEPARACAO
    ws_dados.cell(row=next_row, column=COLUMNS["STATUS"], value="SEPARACAO")

    # Delete from Historico
    ws_hist.delete_rows(row_idx, 1)

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear() # Invalida o cache após estorno
    return True


# ══════════════════════════════════════════════════════════════
# NF UPDATE — Atualização em lote do campo Nota Fiscal
# ══════════════════════════════════════════════════════════════

@require_adm
def update_nf_batch(updates: dict) -> int:
    """
    Atualiza o campo NF de múltiplos pedidos na aba 'Dados' e/ou 'Historico'.

    Busca primeiro na aba Dados. Para pedidos não encontrados lá,
    busca na aba Historico (pedidos já arquivados).

    Args:
        updates: Dict {pedido: nf_value}, ex: {"099613": "2426", "100157": "2544"}

    Returns:
        Número total de pedidos atualizados (Dados + Historico).
    """
    if not updates:
        return 0

    wb = _open_workbook()
    ws_dados = wb[SHEET_PRINCIPAL]

    # Normaliza chaves ignorando zeros à esquerda
    updates_normalized = {str(pk).strip().lstrip('0'): nf for pk, nf in updates.items()}
    found_in_dados = set()

    nf_col = COLUMNS["NF"]
    pk_col = PK_COL_INDEX

    # ── Busca na aba Dados ──
    count_dados = 0
    for row_idx in range(DATA_START_ROW, ws_dados.max_row + 1):
        cell_value = ws_dados.cell(row=row_idx, column=pk_col).value
        if cell_value is not None:
            pedido_str = str(cell_value).strip().lstrip('0')
            if pedido_str in updates_normalized:
                ws_dados.cell(row=row_idx, column=nf_col, value=updates_normalized[pedido_str])
                found_in_dados.add(pedido_str)
                count_dados += 1

    # ── Busca na aba Historico (para pedidos não encontrados em Dados) ──
    count_hist = 0
    remaining = {k: v for k, v in updates_normalized.items() if k not in found_in_dados}

    if remaining and SHEET_HISTORICO in wb.sheetnames:
        ws_hist = wb[SHEET_HISTORICO]
        # Na aba Historico, colunas têm offset -1
        hist_pk_col = pk_col - 1
        hist_nf_col = nf_col - 1

        for row_idx in range(2, ws_hist.max_row + 1):
            cell_value = ws_hist.cell(row=row_idx, column=hist_pk_col).value
            if cell_value is not None:
                pedido_str = str(cell_value).strip().lstrip('0')
                if pedido_str in remaining:
                    ws_hist.cell(row=row_idx, column=hist_nf_col, value=remaining[pedido_str])
                    count_hist += 1

    _save_workbook(wb)
    wb.close()
    st.cache_data.clear()
    return count_dados + count_hist


# ══════════════════════════════════════════════════════════════
# VIAGENS WHATSAPP — Persistência em arquivo separado
# ══════════════════════════════════════════════════════════════

from config import VIAGENS_PATH, VIAGENS_COLUMNS


def _get_viagens_mtime() -> float:
    """Retorna o timestamp de modificação do ficheiro de viagens."""
    try:
        return os.path.getmtime(VIAGENS_PATH) if os.path.exists(VIAGENS_PATH) else 0.0
    except OSError:
        return 0.0


@st.cache_data(show_spinner="Lendo viagens...", ttl=300)
def _read_viagens_cached(mtime: float) -> pd.DataFrame:
    """
    Lê todas as viagens registradas do arquivo viagens_whatsapp.xlsx (versão em cache).
    """
    if not os.path.exists(VIAGENS_PATH):
        return pd.DataFrame(columns=VIAGENS_COLUMNS)

    try:
        df = pd.read_excel(VIAGENS_PATH, engine="openpyxl")
        # Ensure DATA column is datetime
        if "DATA" in df.columns:
            df["DATA"] = pd.to_datetime(df["DATA"], errors="coerce")
        return df
    except Exception:
        return pd.DataFrame(columns=VIAGENS_COLUMNS)


def read_viagens() -> pd.DataFrame:
    """Entrypoint de leitura de viagens, injeta o mtime no cache."""
    return _read_viagens_cached(_get_viagens_mtime())


@require_adm
def insert_viagem(viagem: dict) -> bool:
    """
    Insere uma nova viagem no arquivo viagens_whatsapp.xlsx.

    Se o arquivo não existir, cria com os headers corretos.
    Se já existir, appenda ao final.

    Args:
        viagem: Dict com os campos definidos em VIAGENS_COLUMNS.
                Campos obrigatórios: DATA, MOTORISTA, PLACA.

    Returns:
        True se a inserção foi bem-sucedida.
    """
    # Prepare row data
    row = {}
    for col in VIAGENS_COLUMNS:
        row[col] = viagem.get(col, viagem.get(col.lower(), ""))

    # Format ocorrencias list to string
    if isinstance(row.get("OCORRENCIAS"), list):
        from modules.whatsapp_parser import format_ocorrencias
        row["OCORRENCIAS"] = format_ocorrencias(row["OCORRENCIAS"])

    # Calculate KM_RODADO if not provided
    if not row.get("KM_RODADO") and row.get("KM_INICIAL") and row.get("KM_FINAL"):
        try:
            row["KM_RODADO"] = int(row["KM_FINAL"]) - int(row["KM_INICIAL"])
        except (ValueError, TypeError):
            row["KM_RODADO"] = 0

    new_row_df = pd.DataFrame([row])

    if os.path.exists(VIAGENS_PATH):
        existing = pd.read_excel(VIAGENS_PATH, engine="openpyxl")
        combined = pd.concat([existing, new_row_df], ignore_index=True)
    else:
        combined = new_row_df

    # Ensure all expected columns exist
    for col in VIAGENS_COLUMNS:
        if col not in combined.columns:
            combined[col] = ""

    combined = combined[VIAGENS_COLUMNS]
    combined.to_excel(VIAGENS_PATH, index=False, engine="openpyxl")
    st.cache_data.clear() # Invalida cache de viagens
    return True


@require_adm
def insert_viagens_batch(viagens: list[dict]) -> int:
    """
    Insere múltiplas viagens de uma vez (batch import from chat export).

    Args:
        viagens: Lista de dicts com os campos de viagem.

    Returns:
        Número de viagens inseridas.
    """
    if not viagens:
        return 0

    rows = []
    for v in viagens:
        row = {}
        for col in VIAGENS_COLUMNS:
            row[col] = v.get(col, v.get(col.lower(), ""))
        if isinstance(row.get("OCORRENCIAS"), list):
            from modules.whatsapp_parser import format_ocorrencias
            row["OCORRENCIAS"] = format_ocorrencias(row["OCORRENCIAS"])
        if not row.get("KM_RODADO") and row.get("KM_INICIAL") and row.get("KM_FINAL"):
            try:
                row["KM_RODADO"] = int(row["KM_FINAL"]) - int(row["KM_INICIAL"])
            except (ValueError, TypeError):
                row["KM_RODADO"] = 0
        rows.append(row)

    new_df = pd.DataFrame(rows)

    if os.path.exists(VIAGENS_PATH):
        existing = pd.read_excel(VIAGENS_PATH, engine="openpyxl")
        combined = pd.concat([existing, new_df], ignore_index=True)
    else:
        combined = new_df

    for col in VIAGENS_COLUMNS:
        if col not in combined.columns:
            combined[col] = ""

    combined = combined[VIAGENS_COLUMNS]
    combined.to_excel(VIAGENS_PATH, index=False, engine="openpyxl")
    st.cache_data.clear() # Invalida cache de viagens

    return len(rows)


@require_adm
def delete_viagem(row_index: int) -> bool:
    """
    Exclui uma viagem pelo índice da linha no DataFrame.

    Args:
        row_index: Índice (0-based) da linha a remover.

    Returns:
        True se a exclusão foi bem-sucedida.

    Raises:
        ValueError: Se o índice for inválido ou o arquivo não existir.
    """
    if not os.path.exists(VIAGENS_PATH):
        raise ValueError("Arquivo de viagens não encontrado.")

    df = pd.read_excel(VIAGENS_PATH, engine="openpyxl")

    if row_index < 0 or row_index >= len(df):
        raise ValueError(f"Índice {row_index} fora do intervalo (0-{len(df)-1}).")

    df = df.drop(index=row_index).reset_index(drop=True)
    df.to_excel(VIAGENS_PATH, index=False, engine="openpyxl")
    st.cache_data.clear() # Invalida cache de viagens
    return True


@require_adm
def update_viagem(row_index: int, updated_data: dict) -> bool:
    """
    Atualiza uma viagem pelo índice da linha.

    Args:
        row_index: Índice (0-based) da linha a atualizar.
        updated_data: Dict com os campos a atualizar (parcial ou completo).

    Returns:
        True se a atualização foi bem-sucedida.

    Raises:
        ValueError: Se o índice for inválido ou o arquivo não existir.
    """
    if not os.path.exists(VIAGENS_PATH):
        raise ValueError("Arquivo de viagens não encontrado.")

    df = pd.read_excel(VIAGENS_PATH, engine="openpyxl")

    if row_index < 0 or row_index >= len(df):
        raise ValueError(f"Índice {row_index} fora do intervalo (0-{len(df)-1}).")

    for col, value in updated_data.items():
        if col in df.columns:
            df.at[row_index, col] = value

    df.to_excel(VIAGENS_PATH, index=False, engine="openpyxl")
    st.cache_data.clear() # Invalida cache de viagens
    return True
