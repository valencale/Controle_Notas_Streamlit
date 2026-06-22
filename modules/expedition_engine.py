"""
expedition_engine.py — Motor de lógica de negócio da Expedição Diária.

Responsável por:
- CRUD de planos de expedição em expedicoes.xlsx
- Leitura de pedidos prontos da Esteira (CONTROLE NOTAS.xlsm)
- Geração de Excel formatado (checklist para impressão)
- Cruzamento NF→Pedido via danfe_parser
- Conferência Expedição vs. Viagem (leitura cruzada)

Arquivo de dados: expedicoes.xlsx
    Aba "Expedições" — uma linha por pedido atribuído a um veículo/data
    Aba "Motoristas" — configuração padrão de motorista por placa
"""

import os
import io
from datetime import datetime, date
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import (
    EXPEDITIONS_PATH,
    EXPEDITION_COLUMNS,
    MOTORISTAS_COLUMNS,
    DEFAULT_DRIVERS,
    FLEET_PLATES,
)


# ══════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════

SHEET_EXPEDITIONS = "Expedições"
SHEET_MOTORISTAS = "Motoristas"

# Status transitions
STATUS_DISPATCHED = "EM ROTA"
STATUS_DELIVERED = "ENTREGUE"

# Styles for Excel export
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_HEADER_FILL = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_VEHICLE_FILL = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
_VEHICLE_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
_SUBTOTAL_FILL = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
_SUBTOTAL_FONT = Font(name="Calibri", size=10, bold=True, color="1e293b")
_TOTAL_FILL = PatternFill(start_color="cbd5e1", end_color="cbd5e1", fill_type="solid")
_TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="0f172a")
_DATA_FONT = Font(name="Calibri", size=10, color="1e293b")
_CHECKBOX_FONT = Font(name="Calibri", size=14, color="64748b")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


# ══════════════════════════════════════════════════════════════
# FILE HELPERS
# ══════════════════════════════════════════════════════════════

def _ensure_file():
    """Garante que o arquivo expedicoes.xlsx existe com as abas corretas."""
    if os.path.exists(EXPEDITIONS_PATH):
        return

    wb = openpyxl.Workbook()
    # Aba Expedições
    ws_exp = wb.active
    ws_exp.title = SHEET_EXPEDITIONS
    for col_idx, col_name in enumerate(EXPEDITION_COLUMNS, 1):
        ws_exp.cell(row=1, column=col_idx, value=col_name)

    # Aba Motoristas
    ws_mot = wb.create_sheet(SHEET_MOTORISTAS)
    for col_idx, col_name in enumerate(MOTORISTAS_COLUMNS, 1):
        ws_mot.cell(row=1, column=col_idx, value=col_name)
    # Preenche com os motoristas padrão
    for row_idx, (placa, motorista) in enumerate(DEFAULT_DRIVERS.items(), 2):
        ws_mot.cell(row=row_idx, column=1, value=placa)
        ws_mot.cell(row=row_idx, column=2, value=motorista)

    wb.save(EXPEDITIONS_PATH)
    wb.close()


def _format_date_br(dt) -> str:
    """Formata data para dd/mm/aaaa."""
    if isinstance(dt, str):
        return dt
    if isinstance(dt, datetime):
        return dt.strftime("%d/%m/%Y")
    if isinstance(dt, date):
        return dt.strftime("%d/%m/%Y")
    return str(dt)


def _parse_date_br(date_str: str) -> Optional[date]:
    """Converte string dd/mm/aaaa para date."""
    if not date_str or not isinstance(date_str, str):
        return None
    try:
        return datetime.strptime(date_str.strip(), "%d/%m/%Y").date()
    except ValueError:
        return None


def _extract_city_from_address(address: str) -> str:
    """
    Extrai cidade/UF do endereço Onfinity.
    Ex: 'AVENIDA OTTO BAUMGART, 500 - LOJA 226 - 2049900 - SAO PAULO/SP'
    → 'SAO PAULO/SP'
    """
    if not address:
        return ""
    # O destino é geralmente o último segmento após " - "
    parts = [p.strip() for p in str(address).split(" - ")]
    
    return _find_city_with_uf(parts) or _find_non_numeric_fallback(parts) or str(address)[:50]

def _find_city_with_uf(parts: list[str]) -> str:
    """Procura segmento com /UF (ex: SAO PAULO/SP)"""
    for part in reversed(parts):
        if "/" in part and len(part.split("/")[-1]) == 2:
            return part
    return ""

def _find_non_numeric_fallback(parts: list[str]) -> str:
    """Fallback: último segmento não-numérico"""
    for part in reversed(parts):
        if part and not part.replace(".", "").isdigit():
            return part
    return ""


def _extract_destination_from_obs(obs: str) -> str:
    """
    Extrai destino do campo OBS da Esteira.
    Padrões suportados são separados em sub-funções.
    """
    if not obs:
        return ""
    obs_clean = str(obs).strip()
    if not obs_clean or obs_clean.lower() == "nan":
        return ""

    obs_upper = obs_clean.upper()

    return (
        _extract_dest_pattern_c(obs_upper)
        or _extract_dest_pattern_b(obs_upper)
        or _extract_dest_pattern_a(obs_upper, obs_clean)
        or ""
    )


def _extract_dest_pattern_c(obs_upper: str) -> str:
    """Pattern C: ENDEREÇO DE ENTREGA → extrai CIDADE/UF"""
    import re
    m_end = re.search(
        r'ENDERE[CÇ]O\s+DE\s+ENTREGA\s*:\s*(.+?)(?:FANTASIA|$)',
        obs_upper,
        re.IGNORECASE,
    )
    if m_end:
        bloco = m_end.group(1).strip().rstrip("-~., ")
        # Procura CIDADE/UF (ex: SAO PAULO/SP)
        m_uf = re.search(r'([A-ZÀ-Ú\s]+/[A-Z]{2})', bloco)
        if m_uf:
            return m_uf.group(1).strip().rstrip(".,")
    return ""


def _extract_dest_pattern_b(obs_upper: str) -> str:
    """Pattern B: ENTREGA <CIDADE> <data>"""
    import re
    # Ex: "ENTREGA TABOAO 11/06", "ENTREGA EMBU 3/6", "ENTREGA CD TABOAO"
    m_ent = re.search(
        r'ENTREGA\s+(?:CD\s+)?([A-ZÀ-ÚÃÕ]+(?:\s+(?:DA|DAS|DE|DO|DOS)\s+[A-ZÀ-ÚÃÕ]+)*)',
        obs_upper,
    )
    if m_ent:
        destino = m_ent.group(1).strip()
        # Limpa: remove se capturou palavras-chave como "FATURA", "FALTA"
        if destino and destino not in ("FATURA", "FAT", "FALTA", "IMEDIATO", "CD"):
            return destino
    return ""


def _extract_dest_pattern_a(obs_upper: str, obs_clean: str) -> str:
    """Pattern A: último segmento após ' - ' (padrão Onfinity)"""
    if not obs_upper.startswith("MCS"):
        return ""
        
    obs_stripped = obs_clean.rstrip("-~. ")
    parts = [p.strip().rstrip("-~. ") for p in obs_stripped.split(" - ")]
    if len(parts) >= 2:
        last = parts[-1].upper()
        # Ignora se parece data, código, ou fatura
        if (
            last
            and not last.replace("/", "").replace(".", "").isdigit()
            and not last.startswith(("MCS", "OC ", "FATURA", "NF", "FAT", "FALTA"))
            and 3 <= len(last) <= 30
        ):
            return last
    return ""



# ══════════════════════════════════════════════════════════════
# READ — Pedidos Prontos (da Esteira)
# ══════════════════════════════════════════════════════════════

def get_ready_orders() -> pd.DataFrame:
    """
    Retorna pedidos da Esteira com status CONCLUIDO ou AGUARDANDO NF.
    Estes são os candidatos para atribuição na expedição.

    Returns:
        DataFrame com colunas da Esteira (PEDIDO, CLIENTE, ENDERECO, etc.)
    """
    from modules.excel_handler import read_principal
    df = read_principal()
    if df.empty:
        return df
    mask = df["STATUS"].isin(["CONCLUIDO", "AGUARDANDO NF"])
    return df[mask].copy().reset_index(drop=True)


# ══════════════════════════════════════════════════════════════
# CRUD — Expedições
# ══════════════════════════════════════════════════════════════

def read_expeditions(target_date: Optional[date] = None) -> pd.DataFrame:
    """
    Lê expedições. Se target_date for fornecida, filtra por essa data.

    Args:
        target_date: Data para filtrar (opcional).

    Returns:
        DataFrame com as expedições.
    """
    _ensure_file()

    try:
        df = pd.read_excel(
            EXPEDITIONS_PATH, sheet_name=SHEET_EXPEDITIONS, engine="openpyxl"
        )
    except Exception:
        return pd.DataFrame(columns=EXPEDITION_COLUMNS)

    if df.empty:
        return pd.DataFrame(columns=EXPEDITION_COLUMNS)

    # Ensure all columns exist
    for col in EXPEDITION_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    if target_date is not None:
        target_str = _format_date_br(target_date)
        # Handle both string and datetime formats
        mask = df["DATA_EXPEDICAO"].apply(
            lambda x: _format_date_br(x) == target_str if pd.notna(x) else False
        )
        df = df[mask].copy()

    return df.reset_index(drop=True)


def save_expedition(items: list[dict]) -> int:
    """
    Salva/atualiza um plano de expedição.
    Remove expedições existentes para a mesma DATA_EXPEDICAO e insere as novas.

    Args:
        items: Lista de dicts com campos de EXPEDITION_COLUMNS.

    Returns:
        Número de registros salvos.
    """
    if not items:
        return 0

    _ensure_file()

    # Read existing
    try:
        df_existing = pd.read_excel(
            EXPEDITIONS_PATH, sheet_name=SHEET_EXPEDITIONS, engine="openpyxl"
        )
    except Exception:
        df_existing = pd.DataFrame(columns=EXPEDITION_COLUMNS)

    # Determine the date of the new expedition
    new_date = items[0].get("DATA_EXPEDICAO", "")
    new_date_str = _format_date_br(new_date)

    # Remove existing entries for this date
    if not df_existing.empty:
        mask = df_existing["DATA_EXPEDICAO"].apply(
            lambda x: _format_date_br(x) != new_date_str if pd.notna(x) else True
        )
        df_existing = df_existing[mask].copy()

    # Build new rows
    new_rows = []
    for item in items:
        row = {}
        for col in EXPEDITION_COLUMNS:
            val = item.get(col, "")
            if col in ("DATA_EXPEDICAO", "CRIADO_EM") and not isinstance(val, str):
                val = _format_date_br(val)
            row[col] = val
        new_rows.append(row)

    df_new = pd.DataFrame(new_rows)
    df_combined = pd.concat([df_existing, df_new], ignore_index=True)

    # Ensure all columns
    for col in EXPEDITION_COLUMNS:
        if col not in df_combined.columns:
            df_combined[col] = ""

    df_combined = df_combined[EXPEDITION_COLUMNS]

    # Save — preserving Motoristas sheet
    _save_with_motoristas(df_combined)

    return len(new_rows)


def append_expedition(items: list[dict]) -> int:
    """
    Adiciona registros de expedição SEM remover existentes.
    Ideal para o fluxo reverso (confirmação via DANFE/relatório)
    onde cada pedido pode ter data diferente.

    Evita duplicatas: se o PEDIDO+NF já existir, pula.

    Args:
        items: Lista de dicts com campos de EXPEDITION_COLUMNS.

    Returns:
        Número de registros efetivamente adicionados.
    """
    if not items:
        return 0

    _ensure_file()

    # Read existing
    try:
        df_existing = pd.read_excel(
            EXPEDITIONS_PATH, sheet_name=SHEET_EXPEDITIONS, engine="openpyxl"
        )
    except Exception:
        df_existing = pd.DataFrame(columns=EXPEDITION_COLUMNS)

    # Build set of existing PEDIDO+NF for dedup
    existing_keys = set()
    if not df_existing.empty:
        for _, row in df_existing.iterrows():
            key = f"{str(row.get('PEDIDO', '')).strip()}_{str(row.get('NF', '')).strip()}"
            existing_keys.add(key)

    # Build new rows, skipping duplicates
    new_rows = []
    for item in items:
        pedido = str(item.get("PEDIDO", "")).strip()
        nf = str(item.get("NF", "")).strip()
        key = f"{pedido}_{nf}"

        if key in existing_keys:
            continue  # Já existe — pula

        row = {}
        for col in EXPEDITION_COLUMNS:
            val = item.get(col, "")
            if col in ("DATA_EXPEDICAO", "CRIADO_EM") and not isinstance(val, str):
                val = _format_date_br(val)
            row[col] = val
        new_rows.append(row)
        existing_keys.add(key)

    if not new_rows:
        return 0

    df_new = pd.DataFrame(new_rows)
    df_combined = pd.concat([df_existing, df_new], ignore_index=True)

    # Ensure all columns
    for col in EXPEDITION_COLUMNS:
        if col not in df_combined.columns:
            df_combined[col] = ""

    df_combined = df_combined[EXPEDITION_COLUMNS]

    # Save — preserving Motoristas sheet
    _save_with_motoristas(df_combined)

    return len(new_rows)


def delete_expedition(target_date: date) -> bool:
    """
    Remove todas as expedições de uma data específica.

    Args:
        target_date: Data a remover.

    Returns:
        True se algum registro foi removido.
    """
    _ensure_file()

    try:
        df = pd.read_excel(
            EXPEDITIONS_PATH, sheet_name=SHEET_EXPEDITIONS, engine="openpyxl"
        )
    except Exception:
        return False

    if df.empty:
        return False

    target_str = _format_date_br(target_date)
    mask = df["DATA_EXPEDICAO"].apply(
        lambda x: _format_date_br(x) != target_str if pd.notna(x) else True
    )
    df_filtered = df[mask].copy()

    if len(df_filtered) == len(df):
        return False  # Nothing removed

    _save_with_motoristas(df_filtered)
    return True


def _save_with_motoristas(df_expeditions: pd.DataFrame):
    """Salva o DataFrame de expedições preservando a aba Motoristas."""
    # Read existing motoristas
    drivers = get_default_drivers()

    # Write both sheets
    with pd.ExcelWriter(EXPEDITIONS_PATH, engine="openpyxl", mode="w") as writer:
        df_expeditions[EXPEDITION_COLUMNS].to_excel(
            writer, index=False, sheet_name=SHEET_EXPEDITIONS
        )
        df_mot = pd.DataFrame([
            {"PLACA": p, "MOTORISTA_PADRAO": m}
            for p, m in drivers.items()
        ])
        df_mot.to_excel(writer, index=False, sheet_name=SHEET_MOTORISTAS)


# ══════════════════════════════════════════════════════════════
# MOTORISTAS
# ══════════════════════════════════════════════════════════════

def get_default_drivers() -> dict:
    """
    Lê os motoristas padrão da aba Motoristas.

    Returns:
        Dict {placa: motorista_padrao}
    """
    _ensure_file()

    try:
        df = pd.read_excel(
            EXPEDITIONS_PATH, sheet_name=SHEET_MOTORISTAS, engine="openpyxl"
        )
    except Exception:
        return dict(DEFAULT_DRIVERS)

    if df.empty:
        return dict(DEFAULT_DRIVERS)

    result = dict(DEFAULT_DRIVERS)  # Start with defaults
    for _, row in df.iterrows():
        placa = str(row.get("PLACA", "")).strip()
        motorista = str(row.get("MOTORISTA_PADRAO", "")).strip()
        if placa:
            result[placa] = motorista if motorista != "nan" else ""

    return result


def save_default_drivers(drivers: dict) -> bool:
    """
    Salva os motoristas padrão na aba Motoristas.

    Args:
        drivers: Dict {placa: motorista_padrao}

    Returns:
        True se salvou com sucesso.
    """
    _ensure_file()

    # Read existing expeditions
    try:
        df_exp = pd.read_excel(
            EXPEDITIONS_PATH, sheet_name=SHEET_EXPEDITIONS, engine="openpyxl"
        )
    except Exception:
        df_exp = pd.DataFrame(columns=EXPEDITION_COLUMNS)

    # Write both sheets
    with pd.ExcelWriter(EXPEDITIONS_PATH, engine="openpyxl", mode="w") as writer:
        df_exp.to_excel(writer, index=False, sheet_name=SHEET_EXPEDITIONS)
        df_mot = pd.DataFrame([
            {"PLACA": p, "MOTORISTA_PADRAO": m}
            for p, m in drivers.items()
        ])
        df_mot.to_excel(writer, index=False, sheet_name=SHEET_MOTORISTAS)

    return True


# ══════════════════════════════════════════════════════════════
# NF MATCHING
# ══════════════════════════════════════════════════════════════

def match_nf_to_order(nf_data: dict, orders_df: pd.DataFrame) -> Optional[str]:
    """
    Tenta cruzar dados extraídos de uma NF (DANFE) com um pedido do pool.

    Estratégia de match:
    1. Por Pedido Expresso (se extraído do rodapé da NF)
    2. Por cliente (Razão Social do destinatário × CLIENTE da Esteira)
       - Match exato
       - Match parcial (primeira palavra)
    3. Se houver múltiplos matches para cliente, retorna o primeiro.

    Args:
        nf_data: Dict retornado por danfe_parser.extrair_danfe()
        orders_df: DataFrame de pedidos prontos (get_ready_orders())

    Returns:
        Número do pedido ou None se não encontrou match.
    """
    if not nf_data or orders_df.empty:
        return None

    # Priority 1: Match exactly by PEDIDO if available in NF
    pedido_nf = str(nf_data.get("Pedido", nf_data.get("Pedido_Associado", ""))).strip()
    if pedido_nf:
        pedido_nf_pad = pedido_nf.zfill(6)
        mask_pedido = orders_df["PEDIDO"].astype(str).str.strip().str.zfill(6) == pedido_nf_pad
        matches_pedido = orders_df[mask_pedido]
        if len(matches_pedido) >= 1:
            return str(matches_pedido.iloc[0]["PEDIDO"]).strip()

    cliente_nf = str(nf_data.get("Cliente", "")).strip().upper()
    if not cliente_nf:
        return None

    # Try exact match on CLIENTE
    mask = orders_df["CLIENTE"].str.upper().str.strip() == cliente_nf
    matches = orders_df[mask]

    if len(matches) == 1:
        return str(matches.iloc[0]["PEDIDO"]).strip()

    # Try partial match (first word of client name)
    first_word = cliente_nf.split()[0] if cliente_nf else ""
    if first_word and len(first_word) >= 3:
        mask2 = orders_df["CLIENTE"].str.upper().str.contains(first_word, na=False)
        matches2 = orders_df[mask2]
        if len(matches2) == 1:
            return str(matches2.iloc[0]["PEDIDO"]).strip()

    return None


# ══════════════════════════════════════════════════════════════
# CONFERÊNCIA — Expedição vs. Viagem
# ══════════════════════════════════════════════════════════════

def cross_check_expedition_vs_trip(
    target_date: date, placa: str
) -> Optional[dict]:
    """
    Cruza expedição planejada com viagem registrada.

    Args:
        target_date: Data da expedição.
        placa: Placa do veículo.

    Returns:
        Dict com comparativo ou None se não houver dados.
    """
    from modules.excel_handler import read_viagens

    # Expedição planejada
    df_exp = read_expeditions(target_date)
    if df_exp.empty:
        return None

    exp_veiculo = df_exp[df_exp["VEICULO"] == placa]
    if exp_veiculo.empty:
        return None

    pedidos_planejados = len(exp_veiculo)

    # Viagem registrada
    df_viag = read_viagens()
    if df_viag.empty:
        return {
            "placa": placa,
            "data": _format_date_br(target_date),
            "pedidos_planejados": pedidos_planejados,
            "viagem_registrada": False,
        }

    # Match by PLACA + DATA
    df_viag["DATA"] = pd.to_datetime(df_viag["DATA"], errors="coerce")
    mask = (
        (df_viag["PLACA"] == placa) &
        (df_viag["DATA"].dt.date == target_date)
    )
    viagem_match = df_viag[mask]

    if viagem_match.empty:
        return {
            "placa": placa,
            "data": _format_date_br(target_date),
            "pedidos_planejados": pedidos_planejados,
            "viagem_registrada": False,
        }

    v = viagem_match.iloc[0]
    pedidos_relatados = int(v.get("PEDIDOS", 0) or 0)
    entregas_relatadas = int(v.get("ENTREGAS", 0) or 0)
    ocorrencias = str(v.get("OCORRENCIAS", "") or "")
    km_rodado = int(v.get("KM_RODADO", 0) or 0)
    motorista = str(v.get("MOTORISTA", "") or "")

    # Determine status
    if pedidos_planejados == pedidos_relatados:
        status_pedidos = "✅"
    else:
        status_pedidos = "⚠️"

    if pedidos_relatados > 0 and entregas_relatadas == pedidos_relatados:
        status_entregas = "✅"
    elif pedidos_relatados > 0:
        status_entregas = "⚠️"
    else:
        status_entregas = "—"

    return {
        "placa": placa,
        "data": _format_date_br(target_date),
        "motorista": motorista,
        "pedidos_planejados": pedidos_planejados,
        "pedidos_relatados": pedidos_relatados,
        "entregas_relatadas": entregas_relatadas,
        "ocorrencias": ocorrencias,
        "km_rodado": km_rodado,
        "status_pedidos": status_pedidos,
        "status_entregas": status_entregas,
        "viagem_registrada": True,
    }


# ══════════════════════════════════════════════════════════════
# EXCEL EXPORT — Checklist Formatado para Impressão
# ══════════════════════════════════════════════════════════════

def _setup_worksheet_columns(ws) -> list[str]:
    col_widths = {
        "A": 5, "B": 15, "C": 28, "D": 10, "E": 22,
        "F": 14, "G": 11, "H": 7, "I": 12, "J": 5,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    return ["Ord", "Pedido", "Cliente", "NF", "Destino", "Operação", "Peso", "Vol", "Valor", "☐"]

def _write_checklist_titles(ws, current_row: int, target_date: date, total_pedidos: int, veiculos_count: int) -> int:
    title_text = f"EXPEDIÇÃO DIÁRIA — {_format_date_br(target_date)}"
    subtitle_text = f"TOTAL: {total_pedidos} pedidos  |  {veiculos_count} veículos"

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    title_cell = ws.cell(row=current_row, column=1, value=title_text)
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="0f172a")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    subtitle_cell = ws.cell(row=current_row, column=1, value=subtitle_text)
    subtitle_cell.font = Font(name="Calibri", size=11, color="475569")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    subtitle_cell.fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
    current_row += 2  # Includes empty row
    return current_row

def _clean_data_for_checklist(row: pd.Series) -> dict:
    ordem = int(row.get("ORDEM_ENTREGA", 0))
    if ordem == 99:
        ordem = ""
    peso_val = pd.to_numeric(row.get("PESO", 0), errors="coerce")
    peso_val = peso_val if not pd.isna(peso_val) else 0
    vol_val = pd.to_numeric(row.get("VOLUMES", 0), errors="coerce")
    vol_val = int(vol_val) if not pd.isna(vol_val) else 0
    valor_val = pd.to_numeric(row.get("VALOR", row.get("VALOR_NOTA", 0)), errors="coerce")
    valor_val = valor_val if not pd.isna(valor_val) else 0

    nf_val = str(row.get("NF", "")).strip()
    if nf_val.lower() in ("nan", "none", "0.0", "0"):
        nf_val = ""
    elif nf_val.endswith(".0"):
        nf_val = nf_val[:-2]

    destino_val = str(row.get("DESTINO", "")).strip()
    if destino_val.lower() in ("nan", "none", "0.0", "0"):
        destino_val = ""

    op_val = str(row.get("OPERACAO", "")).strip()
    if op_val.lower() in ("nan", "none", "n/a", "0", ""):
        op_val = ""

    pedido_val = str(row.get("PEDIDO", "")).strip()
    if pedido_val.lower() in ("nan", "none", "n/a", ""):
        pedido_val = "INDISPONIVEL"

    return {
        "ordem": ordem, "pedido": pedido_val, "cliente": str(row.get("CLIENTE", "")),
        "nf": nf_val, "destino": destino_val, "op": op_val, "peso": peso_val,
        "vol": vol_val, "valor": valor_val
    }

def _write_vehicle_section(ws, current_row: int, placa: str, df_v: pd.DataFrame, checklist_cols: list[str]) -> tuple[int, float, float, float]:
    df_v["ORDEM_ENTREGA"] = pd.to_numeric(df_v["ORDEM_ENTREGA"], errors="coerce").fillna(99)
    df_v = df_v.sort_values("ORDEM_ENTREGA").reset_index(drop=True)

    n_pedidos = len(df_v)
    peso_subtotal = pd.to_numeric(df_v["PESO"], errors="coerce").fillna(0).sum()
    vol_subtotal = pd.to_numeric(df_v["VOLUMES"], errors="coerce").fillna(0).sum()
    valor_subtotal = pd.to_numeric(df_v.get("VALOR", df_v.get("VALOR_NOTA", 0)), errors="coerce").fillna(0).sum()

    motorista = df_v.iloc[0].get("MOTORISTA", "") if not df_v.empty else ""
    vehicle_text = f"🚚 {placa}"
    if motorista and str(motorista) != "nan":
        vehicle_text += f" — {motorista}"
    vehicle_text += f"  ({n_pedidos} pedidos  |  {peso_subtotal:.2f} kg  |  {int(vol_subtotal)} vol  |  R$ {valor_subtotal:,.2f})"

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    v_cell = ws.cell(row=current_row, column=1, value=vehicle_text)
    v_cell.font = _VEHICLE_FONT
    v_cell.fill = _VEHICLE_FILL
    v_cell.alignment = _LEFT
    for col_idx in range(1, 11):
        ws.cell(row=current_row, column=col_idx).fill = _VEHICLE_FILL
        ws.cell(row=current_row, column=col_idx).border = _THIN_BORDER
    current_row += 1

    for col_idx, col_name in enumerate(checklist_cols, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    current_row += 1

    for _, row in df_v.iterrows():
        cdata = _clean_data_for_checklist(row)
        values = [
            cdata["ordem"], cdata["pedido"], cdata["cliente"], cdata["nf"],
            cdata["destino"], cdata["op"], f"{cdata['peso']:.2f} kg",
            str(cdata["vol"]) if cdata["vol"] else "", f"R$ {cdata['valor']:,.2f}", "☐",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            if col_idx in (1, 7, 8, 9, 10):
                cell.alignment = _CENTER
            else:
                cell.alignment = _LEFT
            if col_idx == 10:
                cell.font = _CHECKBOX_FONT
        current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    sub_cell = ws.cell(row=current_row, column=1, value="SUBTOTAL")
    sub_cell.font = _SUBTOTAL_FONT
    sub_cell.fill = _SUBTOTAL_FILL
    sub_cell.alignment = Alignment(horizontal="right", vertical="center")
    for col_idx in range(1, 11):
        ws.cell(row=current_row, column=col_idx).fill = _SUBTOTAL_FILL
        ws.cell(row=current_row, column=col_idx).border = _THIN_BORDER

    ws.cell(row=current_row, column=7, value=f"{peso_subtotal:.2f} kg").font = _SUBTOTAL_FONT
    ws.cell(row=current_row, column=7).fill = _SUBTOTAL_FILL
    ws.cell(row=current_row, column=7).alignment = _CENTER

    ws.cell(row=current_row, column=8, value=str(int(vol_subtotal))).font = _SUBTOTAL_FONT
    ws.cell(row=current_row, column=8).fill = _SUBTOTAL_FILL
    ws.cell(row=current_row, column=8).alignment = _CENTER

    ws.cell(row=current_row, column=9, value=f"R$ {valor_subtotal:,.2f}").font = _SUBTOTAL_FONT
    ws.cell(row=current_row, column=9).fill = _SUBTOTAL_FILL
    ws.cell(row=current_row, column=9).alignment = _CENTER

    current_row += 2
    return current_row, peso_subtotal, vol_subtotal, valor_subtotal

def _write_grand_total(ws, current_row: int, total_pedidos: int, peso_total_geral: float, vol_total_geral: float, valor_total_geral: float) -> int:
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    total_cell = ws.cell(row=current_row, column=1, value=f"TOTAL GERAL: {total_pedidos} pedidos")
    total_cell.font = _TOTAL_FONT
    total_cell.fill = _TOTAL_FILL
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    for col_idx in range(1, 11):
        ws.cell(row=current_row, column=col_idx).fill = _TOTAL_FILL
        ws.cell(row=current_row, column=col_idx).border = _THIN_BORDER

    ws.cell(row=current_row, column=7, value=f"{peso_total_geral:.2f} kg").font = _TOTAL_FONT
    ws.cell(row=current_row, column=7).fill = _TOTAL_FILL
    ws.cell(row=current_row, column=7).alignment = _CENTER

    ws.cell(row=current_row, column=8, value=str(int(vol_total_geral))).font = _TOTAL_FONT
    ws.cell(row=current_row, column=8).fill = _TOTAL_FILL
    ws.cell(row=current_row, column=8).alignment = _CENTER

    ws.cell(row=current_row, column=9, value=f"R$ {valor_total_geral:,.2f}").font = _TOTAL_FONT
    ws.cell(row=current_row, column=9).fill = _TOTAL_FILL
    ws.cell(row=current_row, column=9).alignment = _CENTER
    
    return current_row + 1

def export_checklist_excel(target_date: date, df_override: pd.DataFrame = None) -> bytes:
    """
    Gera um Excel formatado como checklist para impressão.

    Layout:
    - Cabeçalho com data e total geral
    - Seção por veículo (com subtotal)
    - Colunas: Ord | Pedido | Cliente | NF | Destino | Peso | Vol | ☐
    - Formatação para A4 paisagem

    Args:
        target_date: Data da expedição.
        df_override: DataFrame a ser usado em vez de recarregar do disco.

    Returns:
        Bytes do arquivo Excel pronto para download.
    """
    if df_override is not None:
        df = df_override.copy()
    else:
        df = read_expeditions(target_date)
        
    if df.empty:
        raise ValueError(f"Nenhuma expedição encontrada para {_format_date_br(target_date)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist"

    checklist_cols = _setup_worksheet_columns(ws)
    current_row = 1
    total_pedidos = len(df)
    veiculos_count = df["VEICULO"].nunique()

    current_row = _write_checklist_titles(ws, current_row, target_date, total_pedidos, veiculos_count)

    peso_total_geral = 0
    vol_total_geral = 0
    valor_total_geral = 0

    veiculos_unicos = df["VEICULO"].unique().tolist()
    ordered_plates = [p for p in FLEET_PLATES if p in veiculos_unicos]
    other_plates = [p for p in veiculos_unicos if p not in FLEET_PLATES]
    all_plates = ordered_plates + other_plates

    for placa in all_plates:
        df_v = df[df["VEICULO"] == placa].copy()
        if df_v.empty:
            continue
        
        current_row, p_sub, v_sub, val_sub = _write_vehicle_section(ws, current_row, placa, df_v, checklist_cols)
        peso_total_geral += p_sub
        vol_total_geral += v_sub
        valor_total_geral += val_sub

    current_row = _write_grand_total(ws, current_row, total_pedidos, peso_total_geral, vol_total_geral, valor_total_geral)

    # ── Print Setup (A4 Landscape) ──
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:J{current_row - 1}"

    # ── Row height for readability ──
    for r in range(1, current_row):
        ws.row_dimensions[r].height = 20

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
# AVAILABLE EXPEDITION DATES
# ══════════════════════════════════════════════════════════════

def get_expedition_dates() -> list[date]:
    """Retorna lista de datas que possuem expedições salvas."""
    _ensure_file()

    try:
        df = pd.read_excel(
            EXPEDITIONS_PATH, sheet_name=SHEET_EXPEDITIONS, engine="openpyxl"
        )
    except Exception:
        return []

    if df.empty or "DATA_EXPEDICAO" not in df.columns:
        return []

    dates = set()
    for val in df["DATA_EXPEDICAO"].dropna().unique():
        parsed = _parse_date_br(str(val))
        if parsed:
            dates.add(parsed)

    return sorted(dates, reverse=True)


# ══════════════════════════════════════════════════════════════
# STATUS TRANSITIONS — Esteira
# ══════════════════════════════════════════════════════════════

def update_orders_status(pedidos: list[str], new_status: str) -> int:
    """
    Atualiza o status de múltiplos pedidos na Esteira (CONTROLE NOTAS.xlsm).

    Transições esperadas:
        - Ao salvar expedição: CONCLUIDO/AGUARDANDO NF → EM ROTA
        - Após conferência OK: EM ROTA → ENTREGUE

    Args:
        pedidos: Lista de números de pedido.
        new_status: Novo status a aplicar.

    Returns:
        Número de pedidos atualizados.
    """
    from modules.excel_handler import update_status_batch
    updates = {str(p).strip(): new_status for p in pedidos}
    return update_status_batch(updates)


def mark_as_dispatched(target_date: date) -> int:
    """
    Marca todos os pedidos de uma expedição como EM ROTA.

    Args:
        target_date: Data da expedição.

    Returns:
        Número de pedidos atualizados.
    """
    df = read_expeditions(target_date)
    if df.empty:
        return 0
    pedidos = df["PEDIDO"].dropna().unique().tolist()
    return update_orders_status(pedidos, STATUS_DISPATCHED)


def mark_as_delivered(target_date: date, placa: str) -> int:
    """
    Marca pedidos de um veículo/data como ENTREGUE após conferência OK.

    Tenta atualizar na aba Dados (principal) primeiro. Se não encontrar
    (pedidos já movidos para Histórico via "Arquivar EM ROTA"), tenta
    atualizar na aba Historico.

    Args:
        target_date: Data da expedição.
        placa: Placa do veículo.

    Returns:
        Número de pedidos atualizados (soma de Dados + Historico).
    """
    df = read_expeditions(target_date)
    if df.empty:
        return 0
    df_v = df[df["VEICULO"] == placa]
    if df_v.empty:
        return 0
    pedidos = df_v["PEDIDO"].dropna().unique().tolist()

    # Tenta atualizar na aba Dados (principal)
    n_dados = update_orders_status(pedidos, STATUS_DELIVERED)

    # Se não atualizou nenhum (ou atualizou menos do que esperado),
    # tenta na aba Historico (pedidos já arquivados)
    if n_dados < len(pedidos):
        from modules.excel_handler import update_status_in_historico
        updates_hist = {str(p).strip(): STATUS_DELIVERED for p in pedidos}
        n_hist = update_status_in_historico(updates_hist)
        return n_dados + n_hist

    return n_dados


def extract_destination(address: str = "", cliente: str = "", obs: str = "") -> str:
    """
    Extrai destino para a expedição.

    Prioridade:
    1. Campo OBS (funciona para TODOS os clientes — BRF, MOGIANA, APETIT, etc.)
    2. Campo ENDERECO (fallback)

    Args:
        address: Campo ENDERECO da Esteira.
        cliente: Campo CLIENTE da Esteira.
        obs: Campo OBS da Esteira.

    Returns:
        String com o destino ou vazio.
    """
    # Tenta extrair do OBS (todos os clientes)
    if obs:
        dest = _extract_destination_from_obs(obs)
        if dest:
            return dest
    # Fallback: extrair do endereço
    return _extract_city_from_address(address)

