"""
historico_backfill.py — Retropreenchimento de DATA_FATURA, DATA_ENTREGA e ENDERECO
na aba Historico a partir do campo OBS existente.
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from config import EXCEL_PATH, SHEET_HISTORICO
from modules.obs_date_extractor import extract_dates_from_obs

# Import address extractor from pdf_parser
try:
    from modules.pdf_parser import _extract_delivery_address_from_obs, _normalize_location
    HAS_ADDRESS_EXTRACTOR = True
except ImportError:
    HAS_ADDRESS_EXTRACTOR = False


def backfill_historico(progress_callback=None):
    """
    Percorre todas as linhas do Historico e preenche DATA_FATURA,
    DATA_ENTREGA e ENDERECO a partir da coluna OBS.

    Historico column layout (1-indexed, offset -1 from Dados):
        A=DATA, B=CLIENTE, C=PEDIDO, D=EMPRESA, E=NF, F=STATUS, G=OBS, H=ENDERECO
        I=DATA_FATURA, J=DATA_ENTREGA

    Args:
        progress_callback: Optional callable(current, total) for progress bar.

    Returns:
        Dict com estatísticas: {"total": N, "fatura_filled": N, "entrega_filled": N, "endereco_filled": N}
    """
    wb = load_workbook(EXCEL_PATH, keep_vba=True)

    if SHEET_HISTORICO not in wb.sheetnames:
        wb.close()
        return {"total": 0, "fatura_filled": 0, "entrega_filled": 0, "endereco_filled": 0}

    ws = wb[SHEET_HISTORICO]
    stats = {"total": 0, "fatura_filled": 0, "entrega_filled": 0, "endereco_filled": 0}

    # Historico: Row 1 = headers, Row 2+ = data
    # Columns: A=1(DATA), B=2(CLIENTE), C=3(PEDIDO), D=4(EMPRESA),
    #          E=5(NF), F=6(STATUS), G=7(OBS), H=8(ENDERECO),
    #          I=9(DATA_FATURA), J=10(DATA_ENTREGA)
    COL_OBS = 7
    COL_ENDERECO = 8
    COL_DATA_FATURA = 9
    COL_DATA_ENTREGA = 10

    total_rows = ws.max_row - 1  # Exclude header
    if total_rows <= 0:
        wb.close()
        return stats

    stats["total"] = total_rows

    for row_idx in range(2, ws.max_row + 1):
        obs_val = ws.cell(row=row_idx, column=COL_OBS).value
        if not obs_val:
            if progress_callback:
                progress_callback(row_idx - 1, total_rows)
            continue

        obs_text = str(obs_val).strip()
        if not obs_text:
            if progress_callback:
                progress_callback(row_idx - 1, total_rows)
            continue

        # Extrair datas
        dates = extract_dates_from_obs(obs_text)

        current_fat = ws.cell(row=row_idx, column=COL_DATA_FATURA).value
        current_ent = ws.cell(row=row_idx, column=COL_DATA_ENTREGA).value

        if dates["DATA FATURA"] and not current_fat:
            ws.cell(row=row_idx, column=COL_DATA_FATURA, value=dates["DATA FATURA"])
            ws.cell(row=row_idx, column=COL_DATA_FATURA).number_format = "DD/MM/YYYY"
            stats["fatura_filled"] += 1

        if dates["DATA ENTREGA"] and not current_ent:
            ws.cell(row=row_idx, column=COL_DATA_ENTREGA, value=dates["DATA ENTREGA"])
            ws.cell(row=row_idx, column=COL_DATA_ENTREGA).number_format = "DD/MM/YYYY"
            stats["entrega_filled"] += 1

        # Extrair endereço se vazio
        current_endereco = ws.cell(row=row_idx, column=COL_ENDERECO).value
        if not current_endereco and HAS_ADDRESS_EXTRACTOR:
            endereco = _extract_delivery_address_from_obs(obs_text)
            if endereco:
                endereco = _normalize_location(endereco)
                ws.cell(row=row_idx, column=COL_ENDERECO, value=endereco)
                stats["endereco_filled"] += 1

        if progress_callback:
            progress_callback(row_idx - 1, total_rows)

    wb.save(EXCEL_PATH)
    wb.close()

    return stats
