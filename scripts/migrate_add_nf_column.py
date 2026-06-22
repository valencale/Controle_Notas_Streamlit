"""
migrate_add_nf_column.py — Script de migração one-shot.

Adiciona o header "NF" na coluna I da planilha CONTROLE NOTAS.xlsm:
    - Aba "Dados": célula I10 (header row)
    - Aba "Historico": coluna H row 1 (se existir)

Cria backup antes de modificar.
"""

import os
import shutil
from datetime import datetime

import openpyxl

# Paths
EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "CONTROLE NOTAS.xlsm")
BACKUP_DIR = os.path.dirname(EXCEL_PATH)


def migrate():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Arquivo não encontrado: {EXCEL_PATH}")
        return False

    # Backup
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"CONTROLE NOTAS_backup_{timestamp}.xlsm")
    shutil.copy2(EXCEL_PATH, backup_path)
    print(f"[OK] Backup criado: {backup_path}")

    # Open workbook
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)

    # -- Aba "Dados" --
    if "Dados" in wb.sheetnames:
        ws = wb["Dados"]
        current_val = ws.cell(row=10, column=9).value  # I10
        if current_val and str(current_val).strip().upper() == "NF":
            print("[INFO] Aba 'Dados': header NF ja existe em I10.")
        else:
            ws.cell(row=10, column=9, value="NF")
            print("[OK] Aba 'Dados': header 'NF' adicionado em I10 (coluna I).")
    else:
        print("[WARN] Aba 'Dados' nao encontrada.")

    # -- Aba "Historico" --
    if "Historico" in wb.sheetnames:
        ws_hist = wb["Historico"]
        # Na aba Historico, as colunas usam offset -1 (A=DATA, B=CLIENTE, etc.)
        # NF seria coluna 8 (H) no Historico (col_idx 9 - 1 = 8)
        current_val_hist = ws_hist.cell(row=1, column=8).value
        if current_val_hist and str(current_val_hist).strip().upper() == "NF":
            print("[INFO] Aba 'Historico': header NF ja existe em H1.")
        else:
            ws_hist.cell(row=1, column=8, value="NF")
            print("[OK] Aba 'Historico': header 'NF' adicionado em H1.")
    else:
        print("[INFO] Aba 'Historico' nao existe ainda (sera criada pelo app).")

    # Save
    wb.save(EXCEL_PATH)
    wb.close()
    print("[OK] Planilha salva com sucesso!")
    return True


if __name__ == "__main__":
    print("=" * 60)
    print("MIGRAÇÃO: Adição da coluna NF à CONTROLE NOTAS.xlsm")
    print("=" * 60)
    migrate()
    print("=" * 60)
